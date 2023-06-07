# -*- coding: utf-8 -*-
"""
@Time ： 2023/6/8 0:56
@Auth ： liangya
@File ：operationExcel.py
"""
from openpyxl import load_workbook

# 读取 Excel 文件中的数据
def get_data_from_excel(file_path, sheet_name):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data