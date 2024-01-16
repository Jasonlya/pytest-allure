# -*- coding: utf-8 -*-
"""
@Time ： 2024/1/16 21:36
@Auth ： liangya
@File ：excel_allure.py
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import pytest
import os


def extract_allure_annotations():
    # 创建 Excel 文档
    workbook = Workbook()
    sheet = workbook.active

    # 设置标题行
    sheet['A1'] = 'Feature'
    sheet['B1'] = 'Story'
    sheet['C1'] = 'Title'

    # 设置标题行的字体为粗体
    for column in range(1, 4):
        cell = sheet.cell(row=1, column=column)
        cell.font = Font(bold=True)

    # 获取 Allure 注解信息
    test_cases = pytest.main(["-vs", "--alluredir", "outputs/allure/temp_jsonreport", "--clean-alluredir"])
    os.system("allure generate outputs/allure/temp_jsonreport -o outputs/allure/html --clean")
    allure_report = openpyxl.load_workbook('outputs/test-cases.xlsx')
    sheet = allure_report.active

    # 从 Allure 报告中提取注解内容并写入 Excel 文档
    for row in sheet.iter_rows(min_row=2, values_only=True):
        feature, story, title = row
        sheet.append([feature, story, title])

    # 保存 Excel 文档
    workbook.save('test-cases.xlsx')


if __name__ == '__main__':
    extract_allure_annotations()
